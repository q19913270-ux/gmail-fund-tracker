"""
Gmail Fund Tracker
目標：從 Gmail 抓取元大投信基金收益分配通知書（PDF 附件），解析後輸出 Excel
"""

import base64
import logging
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

import fitz  # pymupdf
import openpyxl
from dotenv import load_dotenv
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl.styles import Alignment, Font, PatternFill
from tenacity import retry, stop_after_attempt, wait_exponential

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("fund_tracker.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

load_dotenv()

# ── 設定 ──────────────────────────────────────────────────────────
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
CREDENTIALS_FILE = "credentials.json"
TOKEN_FILE = "token.json"
OUTPUT_FILE = "基金收益分配彙整.xlsx"
PDF_DIR = Path("pdfs")
ID_NUMBER = os.getenv("ID_NUMBER", "")
GMAIL_QUERY = os.getenv("GMAIL_QUERY", "from:yuanta 收益分配通知書")

COLUMNS = [
    "基金名稱",
    "基金代號",
    "除權息日期",
    "入帳日期",
    "收到通知日期",
    "持有受益單位數",
    "每單位配發金額",
    "配發總金額",
    "預扣稅額",
    "實際入帳金額",
    "配發類型",
    "累積領取金額",
]

PDF_HEADERS = [
    "證券代號",
    "除息日",
    "持有單位數",
    "每受益權單位分配金額",
    "分配金額",
    "幣別",
    "補扣繳稅額",
    "二代健保補充保費",
    "郵/匯費",
    "實付金額",
    "基金名稱",
    "發放日",
    "給付方式",
    "銀行帳號/相關說明",
]


# ── Gmail 認證 ─────────────────────────────────────────────────────
def get_gmail_service():
    if not Path(CREDENTIALS_FILE).exists():
        raise FileNotFoundError(
            f"找不到 {CREDENTIALS_FILE}，請先放入 Google OAuth client credentials。"
        )

    creds = None
    if Path(TOKEN_FILE).exists():
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
        except Exception as exc:
            logger.warning("無法讀取既有 token，將重新授權：%s", exc)
            creds = None
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                logger.info("Token 已自動更新")
            except Exception as exc:
                logger.warning("Token 更新失敗，將重新授權：%s", exc)
                creds = None
        if not creds or not creds.valid:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w", encoding="utf-8") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


# ── 抓取附件 PDF ───────────────────────────────────────────────────
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def _list_messages(service, query, page_token):
    return service.users().messages().list(
        userId="me", q=query, pageToken=page_token
    ).execute()


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def _get_message(service, message_id):
    return service.users().messages().get(
        userId="me", id=message_id, format="full"
    ).execute()


def fetch_pdf_attachments(service):
    """回傳 list of (received_date, pdf_bytes)"""
    PDF_DIR.mkdir(exist_ok=True)
    pdfs = []
    next_page_token = None

    while True:
        results = _list_messages(service, GMAIL_QUERY, next_page_token)
        messages = results.get("messages", [])

        for msg_meta in messages:
            try:
                msg = _get_message(service, msg_meta["id"])
                received_date = _get_received_date(msg)
                for part in _iter_message_parts(msg.get("payload", {})):
                    filename = (part.get("filename") or "").strip()
                    if not filename.lower().endswith(".pdf"):
                        continue
                    pdf_bytes = _extract_pdf_bytes(service, msg_meta["id"], part)
                    if pdf_bytes:
                        _save_pdf(pdf_bytes, received_date, filename)
                        pdfs.append((received_date, pdf_bytes))
            except HttpError as exc:
                logger.error("讀取郵件 %s 失敗：%s", msg_meta.get("id"), exc)
            except Exception as exc:
                logger.error("處理郵件 %s 時發生未預期錯誤：%s", msg_meta.get("id"), exc)

        next_page_token = results.get("nextPageToken")
        if not next_page_token:
            break

    return pdfs


def _save_pdf(pdf_bytes: bytes, received_date: str, original_filename: str):
    """將 PDF 存到 pdfs/ 目錄，檔名為 日期_原始檔名"""
    safe_date = received_date.replace("-", "")
    dest = PDF_DIR / f"{safe_date}_{original_filename}"
    if dest.exists():
        logger.info("PDF 已存在，略過：%s", dest.name)
        return
    dest.write_bytes(pdf_bytes)
    logger.info("已儲存 PDF：%s", dest.name)


def _get_received_date(message: dict) -> str:
    """將 Gmail internalDate 轉為本地日期字串。"""
    epoch_ms = message.get("internalDate")
    if not epoch_ms:
        # TODO: 若 Gmail API 回傳缺少 internalDate，應改成記錄並跳過，而不是默默使用當天日期。
        return datetime.now().strftime("%Y-%m-%d")
    return datetime.fromtimestamp(int(epoch_ms) / 1000).strftime("%Y-%m-%d")


def _iter_message_parts(payload: dict):
    """遞迴展開 Gmail payload，包含單一附件與 multipart。"""
    parts = payload.get("parts") or []
    if parts:
        for part in parts:
            yield from _iter_message_parts(part)
        return
    yield payload


def _extract_pdf_bytes(service, message_id: str, part: dict) -> bytes | None:
    """從 attachmentId 或 inline data 取回 PDF bytes。"""
    body = part.get("body") or {}
    data = body.get("data")

    if not data and body.get("attachmentId"):
        attachment = (
            service.users()
            .messages()
            .attachments()
            .get(userId="me", messageId=message_id, id=body["attachmentId"])
            .execute()
        )
        data = attachment.get("data")

    if not data:
        logger.warning("郵件 %s 的 PDF 附件缺少內容，已跳過", message_id)
        return None

    try:
        return base64.urlsafe_b64decode(_pad_base64(data))
    except Exception as exc:
        logger.error("郵件 %s 的 PDF 附件解碼失敗：%s", message_id, exc)
        return None


def _pad_base64(data: str) -> str:
    """補齊 Gmail attachment base64 padding。"""
    return data + "=" * (-len(data) % 4)


# ── 解析 PDF ───────────────────────────────────────────────────────
def parse_pdf(pdf_bytes: bytes, received_date: str, password: str) -> dict | None:
    """解析單一 PDF，回傳欄位 dict；失敗回傳 None"""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            if doc.needs_pass and not doc.authenticate(password):
                logger.warning("PDF 密碼錯誤，跳過（%s）", received_date)
                return None
            text = "\n".join(page.get_text() for page in doc)
    except Exception as e:
        logger.error("無法解析 PDF：%s", e)
        return None

    return _extract_fields(text, received_date)


def _roc_to_ce(date_str: str) -> str:
    """民國年轉西元，例如 115/03/20 → 2026-03-20"""
    m = re.match(r"(\d+)/(\d+)/(\d+)", date_str)
    if m:
        return f"{int(m.group(1)) + 1911}-{m.group(2)}-{m.group(3)}"
    return date_str


def _extract_fields(text: str, received_date: str) -> dict:
    """
    元大 PDF 格式：表格標題列在上方，數值列在下方，順序對應。
    標題順序：證券代號、除息日、持有單位數、每受益權單位分配金額、
              分配金額、幣別、補扣繳稅額、二代健保補充保費、
              郵/匯費、實付金額、基金名稱、發放日、給付方式、銀行帳號/相關說明
    """
    lines = [_normalize_line(line) for line in text.split("\n")]
    lines = [line for line in lines if line]

    header_positions = _find_header_positions(lines, PDF_HEADERS)
    if header_positions is None:
        logger.warning("PDF 內容不符合預期格式，跳過（%s）", received_date)
        return None

    values = _extract_table_values(lines, header_positions, PDF_HEADERS)
    if values is None:
        logger.warning("PDF 欄位數不足，跳過（%s）", received_date)
        return None

    # TODO: 若券商 PDF 版面改成多欄或換頁，需改用座標或更穩定的標題配對解析。
    return {
        "基金名稱": values.get("基金名稱", ""),
        "基金代號": values.get("證券代號", ""),
        "除權息日期": _roc_to_ce(values.get("除息日", "")),
        "入帳日期": _roc_to_ce(values.get("發放日", "")),
        "收到通知日期": received_date,
        "持有受益單位數": values.get("持有單位數", "0").replace(",", ""),
        "每單位配發金額": values.get("每受益權單位分配金額", "0").replace(",", ""),
        "配發總金額": values.get("分配金額", "0").replace(",", ""),
        "預扣稅額": values.get("補扣繳稅額", "0").replace(",", ""),
        "實際入帳金額": values.get("實付金額", "0").replace(",", ""),
        "配發類型": values.get("給付方式", ""),
        "累積領取金額": "",  # PDF 中無此欄位，需手動累加
    }


def _normalize_line(line: str) -> str:
    return re.sub(r"\s+", "", line.strip())


def _find_header_positions(lines: list[str], headers: list[str]) -> list[int] | None:
    positions: list[int] = []
    search_from = 0

    for header in headers:
        for idx in range(search_from, len(lines)):
            if lines[idx] == header:
                positions.append(idx)
                search_from = idx + 1
                break
        else:
            return None

    return positions


def _extract_table_values(
    lines: list[str], header_positions: list[int], headers: list[str]
) -> dict[str, str] | None:
    value_lines = lines[header_positions[-1] + 1 :]
    if len(value_lines) < len(headers):
        return None

    extracted = value_lines[: len(headers)]
    return dict(zip(headers, extracted))


# ── 輸出 Excel ────────────────────────────────────────────────────
def save_excel(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基金收益分配"

    # 標題列樣式
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 資料列
    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            # TODO: 若後續需要在 Excel 內做加總/篩選，應將金額欄位轉成數值格式而非字串。
            ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))

    # 自動欄寬
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output_dir = Path(output_path).resolve().parent
    output_dir.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx", dir=output_dir
    ) as tmp_file:
        temp_path = Path(tmp_file.name)

    try:
        wb.save(temp_path)
        temp_path.replace(output_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)

    logger.info("已輸出：%s（共 %d 筆）", output_path, len(records))


# ── 主程式 ────────────────────────────────────────────────────────
def main():
    if not ID_NUMBER:
        logger.error("請在 .env 設定 ID_NUMBER（身分證字號）")
        return

    try:
        logger.info("連線 Gmail...")
        service = get_gmail_service()
    except Exception as exc:
        logger.error("Gmail 認證失敗：%s", exc)
        return

    logger.info("搜尋郵件：%s", GMAIL_QUERY)
    try:
        pdfs = fetch_pdf_attachments(service)
    except Exception as exc:
        logger.error("抓取 Gmail 附件失敗：%s", exc)
        return
    logger.info("找到 %d 個 PDF 附件", len(pdfs))

    records = []
    for received_date, pdf_bytes in pdfs:
        record = parse_pdf(pdf_bytes, received_date, ID_NUMBER)
        if record:
            records.append(record)

    if not records:
        logger.warning("沒有成功解析的記錄")
        return

    save_excel(records, OUTPUT_FILE)


if __name__ == "__main__":
    main()
